import logging
import os.path
import sys
from datetime import datetime, timedelta

import click
import openpyxl
from click_default_group import DefaultGroup
from tqdm import tqdm

from driver import get_chromedriver
from mutually_exclusive_option import MutuallyExclusiveOption
from scrapping.scraper import Scrapper
from config import read_config

logger = logging.getLogger('TwitterFeedScrapper')


@click.group(cls=DefaultGroup, default='foo', default_if_no_args=True)
def cli():
    pass


def manual_login():
    logger.setLevel(logging.INFO)
    setup_file_logger(logger)
    setup_console_logger(logger)
    driver = get_chromedriver(use_proxy=False, headless=False)
    driver.get('https://twitter.com/login')
    input()


@cli.command()
@click.option("-c", "--config", default=os.path.join(os.path.abspath(os.path.dirname(__file__)), 'config.ini'))
@click.option("-o", "--output", required=True)
@click.option("--last-hours", cls=MutuallyExclusiveOption, mutually_exclusive=['date', 'from-date', 'to-date'])
@click.option("--date", cls=MutuallyExclusiveOption, mutually_exclusive=['last-hours', 'from-date', 'to-date'])
@click.option("--from-date", cls=MutuallyExclusiveOption, mutually_exclusive=['last-hours', 'date'])
@click.option("--to-date", cls=MutuallyExclusiveOption, mutually_exclusive=['last-hours', 'date'])
@click.option('--headless', is_flag=True, default=False)
@click.option('--use-proxy', is_flag=True, default=False)
def main(**kwargs):
    logger.setLevel(logging.INFO)
    setup_file_logger(logger)
    setup_console_logger(logger)

    if not (config := read_config(kwargs['config'])):
        return

    tweeted_after, tweeted_before = get_time_range(kwargs['last_hours'], kwargs['date'], kwargs['from_date'],
                                                   kwargs['to_date'])
    logger.info("Запускаем браузер")
    driver = get_chromedriver(use_proxy=kwargs['use_proxy'], headless=kwargs['headless'])
    logger.info("Собираем твиты")
    data = None
    try:
        scrapper = Scrapper(driver, config, tweeted_after, tweeted_before)
        data = scrapper.scrape()
    except Exception as e:
        logger.exception(e)
    finally:
        driver.close()
        driver.quit()
    if data:
        logger.info("Записываем результат")
        write_output(kwargs['output'], data)


def read_input(input_file: str) -> list[str]:
    with open(input_file, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    return [line[:-1] if line[-1] == '\n' else line for line in lines]


def split_list(input_list: list, chunks: int) -> list[list]:
    result = [[] for _ in range(chunks)]
    input_idx = 0
    while input_idx < len(input_list):
        for i in range(chunks):
            if input_idx == len(input_list):
                break
            result[i].append(input_list[input_idx])
            input_idx += 1
    return result


def get_time_range(last_hours, single_date, from_date, to_date) -> tuple[datetime | None, datetime | None]:
    tweeted_before = None
    tweeted_after = None
    if last_hours:
        tweeted_after = datetime.utcnow() - timedelta(hours=int(last_hours))
    if from_date:
        tweeted_after = datetime.strptime(from_date, '%d.%m.%y')
    if to_date:
        to_date = datetime.strptime(to_date, '%d.%m.%y')
        tweeted_before = to_date + timedelta(days=1)
    if single_date:
        tweeted_after = datetime.strptime(single_date, '%d.%m.%y')
        tweeted_before = tweeted_after + timedelta(days=1)
    return tweeted_after, tweeted_before


def write_output(output_file: str, data):
    if os.path.exists(output_file):
        os.remove(output_file)
    fields = ['username', 'datetime', 'content', 'likes', 'retweets']
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i, column_header in enumerate(fields):
        sheet.cell(row=1, column=i + 1).value = column_header
    row_idx = 2
    for _, tweet in tqdm(data.items()):
        if tweet['is_retweet']:
            continue
        for col_idx, value in enumerate(fields):
            sheet.cell(row=row_idx, column=col_idx + 1).value = tweet[value]
        row_idx += 1
    wb.save(output_file)


def setup_file_logger(logger):
    ch = logging.FileHandler('TwitterScrapper.log', encoding='utf-8')
    ch.setLevel(logging.ERROR)
    ch_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s - %(message)s')
    ch.setFormatter(ch_formatter)
    logger.addHandler(ch)


def setup_console_logger(logger):
    format = logging.Formatter(
        "%(asctime)s - %(name)s - %(levelname)s - %(message)s")
    ch = logging.StreamHandler()
    ch.setFormatter(format)
    logger.addHandler(ch)


if __name__ == '__main__':
    if sys.argv[1] == 'manual_login':
        manual_login()
    else:
        main()
