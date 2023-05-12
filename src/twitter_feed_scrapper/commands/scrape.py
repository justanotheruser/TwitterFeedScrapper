import logging
import os.path
from datetime import datetime, timedelta
from typing import Optional

import openpyxl
from tqdm import tqdm

from twitter_feed_scrapper.config import read_config
from twitter_feed_scrapper.driver import get_chromedriver
from twitter_feed_scrapper.web.scraper import Scrapper
from twitter_feed_scrapper.config import Credentials

logger = logging.getLogger('TwitterFeedScrapper')


def scrape_command(config_path: str, output_path: str, last_hours: Optional[str], single_date: Optional[str],
                   from_date: Optional[str], to_date: Optional[str], headless: bool, use_proxy: bool):
    if not (config := read_config(config_path)):
        return
    tweeted_after, tweeted_before = get_time_range(last_hours, single_date, from_date, to_date)
    data = {}
    for credentials in tqdm(config.credentials):
        try:
            account_data = scrape_from_account(credentials, headless, use_proxy, tweeted_after, tweeted_before)
            data.update(account_data)
        except Exception as e:
            logger.exception(e)
    if data:
        logger.info("Записываем результат")
        write_output(output_path, data)


def scrape_from_account(credentials: Credentials, headless: bool, use_proxy: bool, tweeted_after: datetime | None,
                        tweeted_before: datetime | None):
    logger.info("Запускаем браузер")
    driver = get_chromedriver(use_proxy=use_proxy, headless=headless)
    logger.info("Собираем твиты")
    data = {}
    try:
        scrapper = Scrapper(driver, credentials, tweeted_after, tweeted_before)
        data = scrapper.scrape()
    except Exception as e:
        logger.exception(e)
    finally:
        driver.close()
        driver.quit()
        return data

def get_time_range(last_hours, single_date, from_date, to_date) -> tuple[datetime | None, datetime | None]:
    tweeted_before = None
    tweeted_after = None
    if last_hours:
        tweeted_after = datetime.utcnow() - timedelta(hours=int(last_hours))
    if from_date:
        tweeted_after = datetime.strptime(from_date, '%d.%m.%y')
    if to_date:
        to_date = datetime.strptime(to_date, '%d.%m.%y')
        tweeted_before = to_date + timedelta(days=1)
    if single_date:
        tweeted_after = datetime.strptime(single_date, '%d.%m.%y')
        tweeted_before = tweeted_after + timedelta(days=1)
    return tweeted_after, tweeted_before


def write_output(output_file: str, data):
    if os.path.exists(output_file):
        os.remove(output_file)
    fields = ['username', 'datetime', 'content', 'likes', 'retweets']
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i, column_header in enumerate(fields):
        sheet.cell(row=1, column=i + 1).value = column_header
    row_idx = 2
    for _, tweet in tqdm(data.items()):
        if tweet['is_retweet']:
            continue
        for col_idx, value in enumerate(fields):
            sheet.cell(row=row_idx, column=col_idx + 1).value = tweet[value]
        row_idx += 1
    wb.save(output_file)
